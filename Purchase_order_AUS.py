from Unleashed import Unleashed
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook

#Example input
PurchaseOrder = 'PO-0000'+input('Example -> (PO-0000)3248\nPlease enter the last 4 digits of your purchase order number: ')

#API information
Client = Unleashed.Client('', '')
try:
    OrderData = Client.request_endpoint("PurchaseOrders", 'orderNumber='+PurchaseOrder)
except:
    print('Failed to find purchase order, check your internet connection and that you have typed in the correct purchase order number.')
    exit()
    
#Grabs supplier name
SupplierName = OrderData['Items'][0]['Supplier']['SupplierName']

#Opens excel file
wb = load_workbook(''+SupplierName+'.xlsx')
try:
    sheet = wb[PurchaseOrder]
except:
    wb.create_sheet(PurchaseOrder)
    sheet = wb[PurchaseOrder]


#Grab purchase order data (excluding product data)
CurrentRow = 7
for item in OrderData['Items'][0]['PurchaseOrderLines']:
    sheet.cell(CurrentRow,4, item['OrderQuantity'])
    sheet.cell(CurrentRow,7, item['LineTotal'])
    ProductCode = item['Product']['ProductCode']
    print(ProductCode)
    sheet.cell(CurrentRow,2, ProductCode)

    ProductData = Client.return_items("Products", 'productCode='+ProductCode)
    sheet.cell(CurrentRow,16, ProductData[0]['DefaultSellPrice'])
    try:
        sheet.cell(CurrentRow,1, ProductData[0]['Supplier']['SupplierProductCode'])
    except:
        print('Found blank supplier product code on part'+ProductCode+', skipping...')
    sheet.cell(CurrentRow,3, ProductData[0]['ProductDescription'])
    CurrentRow += 1


#Writes to excel file
wb.save('Costings 2023 - '+SupplierName+'.xlsx')
